#!/usr/bin/env python3
"""
加圧トレーニング 予定表自動作成スクリプト
毎週木曜日13時にClaudeが実行します。

- Firebaseから当週の参加希望データを取得
- ペアルールに従ってスロットを割り当て
- Excelスプレッドシートとして出力
"""

import os
import sys
import json
import requests
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────────
# 設定
# ─────────────────────────────────────────────

# .env.local の読み込み（スクリプトの親ディレクトリを探す）
def load_env(env_path: str):
    env = {}
    if not os.path.exists(env_path):
        return env
    with open(env_path, encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                key, _, val = line.partition('=')
                env[key.strip()] = val.strip().strip('"').strip("'")
    return env

# スクリプトから1つ上のディレクトリ（kaatsu-scheduling/）に .env.local がある想定
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
env = load_env(os.path.join(PROJECT_DIR, '.env.local'))

DATABASE_URL = env.get('NEXT_PUBLIC_FIREBASE_DATABASE_URL', '').rstrip('/')

# 時間枠（firebase.ts の TIME_SLOTS と一致させる）
TIME_SLOTS = ['18:00', '18:30', '19:00', '19:30', '20:00', '20:30']
SLOT_END   = ['18:30', '19:00', '19:30', '20:00', '20:30', '21:00']

# ペアルール（両方いれば必ず同じ時間にする）
PAIR_RULES = [
    ('ともき', 'ゆうか'),
    ('はやと', 'ゆうた'),
    ('けいや', 'そり'),
    ('かいしん', 'ことみ'),
]

# ─────────────────────────────────────────────
# 日付ユーティリティ
# ─────────────────────────────────────────────

def get_week_key(date: datetime = None) -> str:
    """firebase.ts の getWeekKey と同じロジック（火曜日ベース）"""
    d = date or datetime.now()
    day = d.weekday()   # 0=月 1=火 2=水 3=木 ...
    diff = -(day - 1) if day >= 1 else -(day + 6)
    tuesday = d + timedelta(days=diff)
    return tuesday.strftime('%Y-%m-%d')

def get_thursday_label(week_key: str) -> str:
    """ウィークキー → 木曜日の日付ラベル (例: 4月10日（木）)"""
    tuesday = datetime.strptime(week_key, '%Y-%m-%d')
    thursday = tuesday + timedelta(days=2)
    return f"{thursday.month}月{thursday.day}日（木）"

# ─────────────────────────────────────────────
# Firebase データ取得
# ─────────────────────────────────────────────

def fetch_responses(week_key: str) -> dict:
    if not DATABASE_URL:
        print("[ERROR] NEXT_PUBLIC_FIREBASE_DATABASE_URL が設定されていません", file=sys.stderr)
        return {}
    url = f"{DATABASE_URL}/kaatsu/responses/{week_key}.json"
    try:
        r = requests.get(url, timeout=10)
        r.raise_for_status()
        data = r.json()
        return data if isinstance(data, dict) else {}
    except Exception as e:
        print(f"[ERROR] Firebase 取得失敗: {e}", file=sys.stderr)
        return {}

# ─────────────────────────────────────────────
# スケジューリングロジック
# ─────────────────────────────────────────────

def build_schedule(responses: dict) -> dict:
    """
    参加者の回答を元に、各スロットに2人ずつ割り当てる。

    Returns:
        {slot_id: [name1, name2 or None], ...}
    """
    # 参加者ごとの利用可能スロット（○のみ。△は除外）
    avail: dict[str, list[str]] = {}
    for entry in responses.values():
        name = entry.get('name', '').strip()
        slots = entry.get('slots', {})
        ok = [s for s in TIME_SLOTS if slots.get(s) == '○']
        if name and ok:
            avail[name] = ok

    if not avail:
        return {s: [] for s in TIME_SLOTS}

    assigned: dict[str, list[str]] = {s: [] for s in TIME_SLOTS}
    scheduled: set[str] = set()  # 割り当て済みの人

    def common_slot(names: list[str]) -> str | None:
        """複数人が全員○をつけている共通スロットを探す（空きがある枠のみ）"""
        for s in TIME_SLOTS:
            if len(assigned[s]) >= 2:
                continue
            if all(s in avail.get(n, []) for n in names):
                return s
        return None

    def solo_slot(name: str) -> str | None:
        """1人分の○スロットを探す（空きがある枠のみ）"""
        for s in avail.get(name, []):
            if len(assigned[s]) < 2:
                return s
        return None

    used_slots: list[str] = []

    # ── Step 1: ペアルール適用 ──
    for p1, p2 in PAIR_RULES:
        both = p1 in avail and p2 in avail

        if both and p1 not in scheduled and p2 not in scheduled:
            slot = common_slot([p1, p2])
            if slot:
                # 共通の○スロットがある → 同じ枠に入れる
                assigned[slot].extend([p1, p2])
                scheduled.update([p1, p2])
                used_slots.append(slot)
            else:
                # 共通の○スロットがない → それぞれ別々に残りに回す（Step2で処理）
                pass

    # ── Step 2: 残りの参加者をペアに ──
    remaining = [n for n in avail if n not in scheduled]

    # ペアルールに含まれる人で片方だけ残っているケースを先に処理
    rule_members = {n for pair in PAIR_RULES for n in pair}
    rule_remaining = [n for n in remaining if n in rule_members]
    other_remaining = [n for n in remaining if n not in rule_members]
    remaining_ordered = rule_remaining + other_remaining

    while remaining_ordered:
        p1 = remaining_ordered.pop(0)
        if p1 in scheduled:
            continue

        # p1 と○が共通するスロットがある人を探す
        partner = None
        for p2 in remaining_ordered:
            if p2 in scheduled:
                continue
            if common_slot([p1, p2]) is not None:
                partner = p2
                break

        if partner:
            remaining_ordered.remove(partner)
            slot = common_slot([p1, partner])
        else:
            slot = solo_slot(p1)

        if slot:
            names_to_add = [p1] + ([partner] if partner else [])
            for n in names_to_add:
                if len(assigned[slot]) < 2:
                    assigned[slot].append(n)
                    scheduled.add(n)
            if slot not in used_slots:
                used_slots.append(slot)
        else:
            # ○スロットが全て埋まっている場合は未割り当てのまま（スケジュール外）
            scheduled.add(p1)  # ループ終了のため登録だけする

    return assigned

# ─────────────────────────────────────────────
# Excel 出力
# ─────────────────────────────────────────────

def create_spreadsheet(schedule: dict, thursday_label: str, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '予定表'

    # ── スタイル定義 ──
    BLUE_DARK   = '1E40AF'
    BLUE_MID    = '3B82F6'
    BLUE_LIGHT  = 'DBEAFE'
    BLUE_PALE   = 'EFF6FF'
    WHITE       = 'FFFFFF'
    TEXT_DARK   = '1E293B'
    TEXT_NAME   = '1D4ED8'

    def thin_border(color='BFDBFE'):
        s = Side(style='thin', color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def cell_style(cell, value, bold=False, size=11, color=TEXT_DARK,
                   fill_color=None, align='center', border=True):
        cell.value = value
        cell.font = Font(bold=bold, size=size, color=color,
                         name='Meiryo UI')
        if fill_color:
            cell.fill = PatternFill('solid', fgColor=fill_color)
        cell.alignment = Alignment(horizontal=align, vertical='center',
                                   wrap_text=False)
        if border:
            cell.border = thin_border()

    # ── 行1: タイトル ──
    ws.merge_cells('A1:C1')
    cell_style(ws['A1'],
               f'💪 加圧トレーニング予定表　{thursday_label}',
               bold=True, size=13, color=WHITE,
               fill_color=BLUE_DARK)
    ws.row_dimensions[1].height = 32

    # ── 行2: ヘッダー ──
    for col, h in enumerate(['時間', '参加者①', '参加者②'], 1):
        cell_style(ws.cell(row=2, column=col), h,
                   bold=True, size=10, color=WHITE,
                   fill_color=BLUE_MID)
    ws.row_dimensions[2].height = 22

    # ── 行3〜: データ ──
    for i, (slot, end) in enumerate(zip(TIME_SLOTS, SLOT_END)):
        row = i + 3
        people = schedule.get(slot, [])
        p1 = people[0] if len(people) > 0 else ''
        p2 = people[1] if len(people) > 1 else ''
        fill = BLUE_PALE if i % 2 == 0 else BLUE_LIGHT

        cell_style(ws.cell(row=row, column=1),
                   f'{slot}〜{end}',
                   bold=True, size=10, color=BLUE_DARK,
                   fill_color=fill)
        cell_style(ws.cell(row=row, column=2), p1,
                   bold=bool(p1), size=12,
                   color=TEXT_NAME if p1 else TEXT_DARK,
                   fill_color=fill)
        cell_style(ws.cell(row=row, column=3), p2,
                   bold=bool(p2), size=12,
                   color=TEXT_NAME if p2 else TEXT_DARK,
                   fill_color=fill)
        ws.row_dimensions[row].height = 30

    # ── 列幅 ──
    ws.column_dimensions['A'].width = 17
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14

    wb.save(output_path)
    return output_path

# ─────────────────────────────────────────────
# メイン
# ─────────────────────────────────────────────

def main():
    week_key = get_week_key()
    thursday_label = get_thursday_label(week_key)

    print(f"週キー      : {week_key}")
    print(f"木曜日      : {thursday_label}")
    print(f"Firebase URL: {DATABASE_URL or '(未設定)'}")

    responses = fetch_responses(week_key)
    print(f"回答数      : {len(responses)} 件")

    schedule = build_schedule(responses)

    print("─── スケジュール ───")
    for slot, end in zip(TIME_SLOTS, SLOT_END):
        people = schedule.get(slot, [])
        print(f"  {slot}〜{end}  {' / '.join(people) if people else '（空き）'}")

    # 出力先: プロジェクトルート（kaatsu-scheduling/）直下
    filename = f"予定表_{thursday_label.replace('（木）', '')}.xlsx"
    output_path = os.path.join(PROJECT_DIR, filename)

    create_spreadsheet(schedule, thursday_label, output_path)
    print(f"\n✅ 保存完了: {output_path}")
    return output_path

if __name__ == '__main__':
    main()
